#! python3
"""
Uses data from team schedule worksheet and creates calendar events on the google calendar,
which is then synced to the TOC website via Cloversites built-in sync function
"""

from __future__ import print_function
from datetime import timedelta
import datetime as dt
import pickle
import os.path
import sys
import openpyxl
from ics import Calendar, Event
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/calendar']
SCRIPTPATH = os.path.dirname(os.path.realpath(sys.argv[0])) + '\\'

def main():
    # Load excel sheet
    path = '\\\\192.168.86.214\\SharePi\\' # Network drives need double slashes \\ at the top
    fileName = 'TOC Team Schedule Jan-Mar20.xlsm'
    ws = load_workbook(path, fileName)['Web Cal']

    # Get schedule data from excel file
    tocCalData = get_xl_data(ws)

    # TODO Compare
        # if log file doesnt exist, create
        # if it does, load that data
        # compare the two
        # email differences

    # Connect to the Google Calendar API and get calendar names and IDs
    service = check_creds(SCRIPTPATH)
    calids = get_cal_ids(service)
    calNames = ['TOC Test']
    
    # Delete duplicate events
    eventids = get_event_ids(service, calNames, calids, tocCalData)
    delete_events(service, calNames, calids, eventids)

    # Upload calendar data to Google Calendar
    update_gcal(tocCalData, service, calNames, calids)

    # Create .ics file (unnecessary at the moment, but may have a future use)
    # create_ics(tocCalData, 'TOC_cal')


def load_workbook(path, fileName):
    """
    Loads the excel workbook for reading
    """
    print(f'Reading \'{fileName}\'... ', flush=True, end='')
    workbook = openpyxl.load_workbook(path + fileName, read_only=True, data_only=True) 
    #print('Done')
    return workbook

def get_xl_data(sheet):
    """
    Returns list of dictionaries: name, date and description from the worsheet
    """
    #print(f'Parsing calendar data... ', flush=True, end='')
    dates = []
    # Get dates
    for r in range(1, 40, 3):
        dates.append(sheet.cell(row=r, column=1).value)
    descriptions = []
    # Get description data
    for r in range(2, 41, 3):
        description = ''
        for c in range(1, 40):
            description += str(sheet.cell(row=r, column=c).value) + ' '
        descriptions.append(description)
    # Create dictionaries
    sourceCalDataList = []
    sourceCalData = {}
    for i in range(len(dates)):
        name = 'Team Schedule'
        # Add 'Empty Services'
        if dates[i] is None:
            name = 'Empty Service'
            dates[i] = dates[i-1] + timedelta(days=7) # Add the date back in, 7 days after the previous
            descriptions[i] = name
        sourceCalData['name'] = name
        sourceCalData['date'] = dates[i].date()
        sourceCalData['description'] = descriptions[i]
        # Create list of dictionaries
        sourceCalDataList.append(sourceCalData.copy())
    print('Done')
    return sourceCalDataList

def check_creds(credPath):
    """
    Opens Google Calendar API connection
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                credPath + 'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
    service = build('calendar', 'v3', credentials=creds)
    return service

def get_cal_ids(service):
    """
    Return all calendar names (key) and ids (value) associated with the master calendar email address
    """
    page_token = None
    calids = {}
    while True:
        calendar_list = service.calendarList().list(pageToken=page_token).execute()
        for calendar_list_entry in calendar_list['items']:
            calids[calendar_list_entry['summary']] = calendar_list_entry['id']
        page_token = calendar_list.get('nextPageToken')
        if not page_token:
            return calids

def get_event_ids(service, calNames, calids, sourceCalData):
    """ Returns a list of event ids of events in sourceCalData """
    eventids = []
    # Get the first and last date from the source calendar data
    startDate = dt.datetime(sourceCalData[0]['date'].year, sourceCalData[0]['date'].month, sourceCalData[0]['date'].day).isoformat()+'Z'
    endDate = (dt.datetime(sourceCalData[-1]['date'].year, sourceCalData[-1]['date'].month, sourceCalData[-1]['date'].day)+ timedelta(days=1)).isoformat() + 'Z'

    # Store the dates and names from the source calendar for comparison
    sourceDate = []
    sourceName = []
    for sourceData in sourceCalData:
        sourceDate.append(sourceData['date'].strftime('%Y-%m-%d'))
        sourceName.append(sourceData['name'])

    # Get all events within the dates in the source calendar
    for calName in calNames:
        if calName not in calids:
            print(f'Calendar \'{name}\' not found')
            return None
        for cal in calids:
            if  cal == calName:
                calid = calids.get(cal)
                print(f'Collecting event ids from calendar \'{calName}\'... ', end='', flush=True)
                events_result = service.events().list(  calendarId=calid, timeMin=startDate,
                                                        timeMax=endDate, singleEvents=True,
                                                        orderBy='startTime').execute()
                events = events_result.get('items', [])

                # if the event name and date match the source calendar name and date, populate event id
                for eid in events: 
                    idDate = eid['start'].get('date')
                    idName = eid['summary']
                    if idDate in sourceDate and idName in sourceName:
                        eventids.append(eid['id'])
                print('Done')
    return eventids

def delete_events(service, calNames, calids, eventids):
    """ Deletes events based on eventids """
    
    if eventids is None:
        print('No duplicate events found')
        return

    for calName in calNames:
        if calName not in calids:
            print(f'Calendar \'{calName}\' not found, no events to delete')
            return
    
        print('Deleting duplicate events... ', end='', flush=True)
        for cal in calids:
            if cal == calName:
                calid = calids.get(cal)
                for eid in eventids:
                    #print(f'Deleting event id \'{eid}\'...', end='', flush = True)
                    service.events().delete(calendarId=calid, eventId=eid).execute()
                    #print('Done')
    print('Done')
    return

def update_gcal(sourceCalData, service, calNames, calids):
    """
    Uploads event data to the calendars passed
    """
    for calendar in calNames:
        for cal in calids:
            if cal == calendar:
                calid = calids.get(cal)
                print(f'Loading events to calendar \'{cal}\'...')
                for data in sourceCalData:
                    name = data.get('name')
                    startDate = data.get('date').strftime('%Y-%m-%d')
                    endDate = (data.get('date') + timedelta(days=1)).strftime('%Y-%m-%d')
                    description = data.get('description')
                    # Create the event
                    event = {
                        'summary': name,
                        'location': 'The Oregon Community 700 NE Dekum St. Portland, OR 97211',
                        'description': description,
                        'start': {
                            'date': startDate,
                            'timeZone': 'America/Los_Angeles',
                        },
                        'end': {
                            'date': endDate,
                            'timeZone': 'America/Los_Angeles',
                        }
                    }
                    event = service.events().insert(calendarId=calid, body=event).execute()
                    #event = service.events().update(calendarId=calid, body=event).execute()
                    
                    print(f'     Event \'{name}\' created on {startDate}')
                print('Done')
        if calendar not in calids:
            print(f'Calendar \'{calendar}\' not found.')
    return

def create_ics(sourceCalData, outFile):
    """ Creates an .ics file for uploading to calendar services """
    cal = Calendar()
    print(f'Writing ICS file \'{outFile}.ics\'... ', flush=True, end='')
    for data in sourceCalData:
        event = Event()
        event.name = data.get('name')
        event.begin = data.get('date')
        event.make_all_day()
        event.description = data.get('description')
        event.location = 'The Oregon Community 700 NE Dekum St. Portland OR'
        cal.events.add(event)
    with open(outFile + '.ics', 'w', newline='') as f: # Clover calendar wont read ics with extra carriage returns
        f.writelines(cal)
    print('Done')
    return

if __name__ == '__main__':
    main()
